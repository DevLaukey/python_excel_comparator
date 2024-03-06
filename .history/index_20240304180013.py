import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

def read_excel_file(file_path):
    if "overall" in file_path:
        return pd.read_excel(file_path, index_col=0).T
    else:
        return pd.read_excel(file_path, index_col=0)

def compare_and_visualize(scenario_name, df_operations, df_overall, writer):
    relevant_columns = ["Total revenues (€)", "Day-ahead revenues (€)", "Energy lack balancing revenues (€)",
                        "Energy surplus balancing revenues (€)", "primary up band revenues (€)",
                        "primary down band revenues (€)", "secondary up band revenues (€)",
                        "secondary down band revenues (€)"]

    try:
        df_compare = pd.DataFrame({
            "Operations Results": df_operations.loc[:, relevant_columns].sum(),
            "Overall Results": df_overall.loc[:, relevant_columns].sum()
        })

        # Additional statistical calculations
        df_statistics = pd.DataFrame({
            "Mean": df_compare.mean(),
            "Min": df_compare.min(),
            "Max": df_compare.max(),
            "Std Dev": df_compare.std(),
            "Sum": df_compare.sum(),
            "Count": df_compare.count()
        })

        # Plotting bar chart for comparison
        plt.figure(figsize=(10, 6))
        df_compare.plot(kind='bar', rot=45, color=['skyblue', 'lightgreen'])
        plt.title(f"Comparison for {scenario_name}")
        plt.ylabel("Revenue (€)")
        plt.savefig(f"Comparison_{scenario_name}.png")
        plt.close()

        # Save the comparison chart and statistical information to Excel
        sheet_name_compare = f"{scenario_name} Comparison"
        sheet_name_statistics = f"{scenario_name} Statistics"

        if sheet_name_compare in writer.book.sheetnames:
            idx_compare = writer.book.sheetnames.index(sheet_name_compare)
            writer.book.remove(writer.book.worksheets[idx_compare])

        if sheet_name_statistics in writer.book.sheetnames:
            idx_statistics = writer.book.sheetnames.index(sheet_name_statistics)
            writer.book.remove(writer.book.worksheets[idx_statistics])

        df_compare.to_excel(writer, sheet_name=sheet_name_compare, index=True)  # Set index to True
        df_statistics.to_excel(writer, sheet_name=sheet_name_statistics, index=True)  # Set index to True

    except KeyError as e:
        print(f"\nError: One or more relevant columns not found in the dataframes.")
        print(f"Make sure the specified columns are present in both dataframes.")
        print(f"Error details: {e}")

def create_excel_template(scenario_names):
    writer = pd.ExcelWriter('Output_Template.xlsx', engine='openpyxl')
    for scenario_name in scenario_names:
        df_compare = pd.DataFrame()
        df_statistics = pd.DataFrame()
        df_compare.to_excel(writer, sheet_name=f"{scenario_name} Comparison", index=True)  # Set index to True
        df_statistics.to_excel(writer, sheet_name=f"{scenario_name} Statistics", index=True)  # Set index to True
    writer._save()

def main():
    num_scenarios = int(input("Enter the number of scenarios to compare (1, 2, or 3): "))
    scenario_files = []
    scenario_names = []
    for i in range(num_scenarios):
        scenario_name = input(f"Enter the name for scenario {i + 1}: ")
        file_path_operations = input(f"Enter the file path for scenario {i + 1} operations results: ")
        file_path_overall = input(f"Enter the file path for scenario {i + 1} overall results: ")
        scenario_files.append((file_path_operations, file_path_overall))
        scenario_names.append(scenario_name)

    create_excel_template(scenario_names)

    with pd.ExcelWriter('Output_Template.xlsx', engine='openpyxl', mode='a') as writer:
        for i, (file_path_operations, file_path_overall) in enumerate(scenario_files):
            df_operations = read_excel_file(file_path_operations)
            df_overall = read_excel_file(file_path_overall)
            compare_and_visualize(scenario_names[i], df_operations, df_overall, writer)

if __name__ == "__main__":
    main()
